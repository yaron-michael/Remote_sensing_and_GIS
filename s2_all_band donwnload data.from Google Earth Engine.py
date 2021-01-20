#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Oct  1 11:35:47 2019

@author: yaron
"""

import ee
import os
from ipygee import *
import geopandas as gpd
import pandas as pd 
from datetime import datetime
import openpyxl
import xlsxwriter

ee.Initialize()


feature_to_use ='users/yaron1205/DAVID_BONFL/South2020'

id_colum_to_use = 'field'

patch_to_save_file_with_file_name = '/home/yaron/South2020.xlsx'




def donwload_Sentinel_2_Level_1C (Feature_to_use,Column_name,Patch_to_save_file_with_file_name):
    '''
    to add
    the scale of all the band wil be 10m

            Parameters:
                    a (str): the gee feature to use like this 'users/yaron1205/DAVID_BONFL/South250220'
                    b (str): feature column to use  of gee feature 
                    b (str): patch_to_save_file_with_file_name like this in linux '/home/yaron/South250220.xlsx'

            Returns:
                    binary_sum (str): Binary string of the sum of a and b
                    
    '''
    Scale = 0.0001 
    def Scale_function(x):
     x= x*Scale
     return x
    # Create a workbook and add a worksheet.

    workbook = xlsxwriter.Workbook(patch_to_save_file_with_file_name)
    worksheet = workbook.add_worksheet()
    workbook.close()
     
    


    feature = ee.FeatureCollection(feature_to_use)
    fc_temp = feature.getInfo()
    fc_info =fc_temp['features']


    list_of_vale_to_loop =[]

   # Using for loop 
    #count=0
    for i in fc_info: 
   # count +=1
    #print (str(count))
     print(i['properties'][id_colum_to_use]) 
    #temp = i['properties'][id_colum_to_use]
    #x=temp.replace('/', '--')
    #list_of_vale_to_loop.append(x+str(count))
     list_of_vale_to_loop.append(i['properties'][id_colum_to_use])

    list_of_vale_to_loop = list_of_vale_to_loop
  

    list_of_vale_to_loop.sort()
   
    list_of_sheet =[]
    list_of_sheet_name =[]
    count=0
    
   #    start the work 
    
    for vale in  list_of_vale_to_loop:
     count +=1
     vale =str(vale)
     print (vale)
     FC = (ee.FeatureCollection(feature_to_use).filter(ee.Filter().eq(id_colum_to_use, vale)))
     print(FC)  
     col = ee.ImageCollection('COPERNICUS/S2').filterBounds(feature)
     time_series = col.filterDate('2019-11-01', '2020-05-01')
     bands = ['B1', 'B2', 'B3','B4', 'B5', 'B6','B7', 'B8', 'B8A','B9', 'B10', 'B11','B12','QA60']
     chart_ts = chart.Image.series(**{
     'imageCollection': time_series, 
     'region': FC,
     'scale': 10,
     'bands': bands,
     'reducer': ee.Reducer.mean()}) 
    
     print(chart_ts.dataframe)   
 
     dataframe = chart_ts.dataframe.apply(Scale_function, axis=1)
    
     list_of_sheet.append(dataframe)
    #vale=vale.replace('/', '--')
     print (count)
     x = vale.replace('/', '--')
     list_of_sheet_name.append(x+"  "+str(count))
     
    for sheet_data, sheet in zip(list_of_sheet, list_of_sheet_name):
     print(sheet_data, sheet)
    
     sheet_data['DATE'] = sheet_data.index
     sheet_data = sheet_data.reindex(columns=['DATE','B1', 'B2', 'B3','B4', 'B5', 'B6','B7', 'B8', 'B8A','B9', 'B10', 'B11','B12','QA60'])
     sheet_data['DATE'] = sheet_data['DATE'].apply(lambda x: x.strftime("%d/%m/%Y"))
     sheet_data = sheet_data.rename({'QA60': 'CloudMask' } , axis='columns')
     sheet_data['CloudMask'] = sheet_data['CloudMask'].apply(lambda x: 1 if x > 0 else 0)
      
     writer = pd.ExcelWriter(patch_to_save_file_with_file_name, engine='openpyxl')
     book = openpyxl.load_workbook(patch_to_save_file_with_file_name)
     writer.book = book
     writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
     sheet_data.to_excel(writer, sheet_name=sheet, index=False )
     writer.save()
     
     
     
    workbook=openpyxl.load_workbook(patch_to_save_file_with_file_name)
    std=workbook['Sheet1']
    workbook.remove(std)
    workbook.save(patch_to_save_file_with_file_name)
    return  "END"
   
donwload_Sentinel_2_Level_1C(feature_to_use,id_colum_to_use,patch_to_save_file_with_file_name)        
     
    
 




    
    
    













    






